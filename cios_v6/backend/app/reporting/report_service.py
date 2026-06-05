"""
Reporting Service — Multi-format clinical report generation.
Supports: PDF, CSV, Excel (XLSX), XPS (via PDF)
"""
import os
import csv
import uuid
import json
from datetime import datetime
from io import BytesIO, StringIO
from typing import Optional, List
from pathlib import Path
from loguru import logger

from app.core.config import settings


class ReportingService:

    def __init__(self):
        Path(settings.REPORTS_DIR).mkdir(parents=True, exist_ok=True)

    def generate_patient_report(
        self,
        patient: dict,
        vitals: List[dict],
        diagnoses: List[dict],
        labs: List[dict],
        ai_assessment: dict,
        format: str = "pdf",
        generated_by: str = "CIOS System"
    ) -> str:
        """Generate a comprehensive patient report. Returns file path."""
        report_id = str(uuid.uuid4())[:8].upper()
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        patient_name = patient.get("full_name", "Unknown").replace(" ", "_")
        filename = f"CIOS_{patient_name}_{timestamp}_{report_id}"

        if format == "pdf":
            return self._generate_pdf(filename, patient, vitals, diagnoses, labs, ai_assessment, generated_by)
        elif format == "csv":
            return self._generate_csv(filename, patient, vitals, diagnoses, labs, ai_assessment)
        elif format in ("excel", "xlsx"):
            return self._generate_excel(filename, patient, vitals, diagnoses, labs, ai_assessment)
        elif format == "xps":
            # XPS via PDF
            path = self._generate_pdf(filename, patient, vitals, diagnoses, labs, ai_assessment, generated_by)
            return path  # In production, convert PDF to XPS
        else:
            raise ValueError(f"Unsupported format: {format}")

    def _generate_pdf(self, filename, patient, vitals, diagnoses, labs, ai_assessment, generated_by) -> str:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.colors import HexColor, white, black
            from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                             Table, TableStyle, HRFlowable)
            from reportlab.lib.units import cm
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        except ImportError:
            return self._generate_csv(filename, patient, vitals, diagnoses, labs, ai_assessment)

        filepath = os.path.join(settings.REPORTS_DIR, f"{filename}.pdf")

        # Colors
        NAVY = HexColor("#0B1E3D")
        TEAL = HexColor("#0EA5E9")
        CRITICAL = HexColor("#DC2626")
        HIGH = HexColor("#F97316")
        MEDIUM = HexColor("#EAB308")
        LOW = HexColor("#22C55E")
        LIGHT_BG = HexColor("#F8FAFC")

        doc = SimpleDocTemplate(
            filepath, pagesize=A4,
            leftMargin=2*cm, rightMargin=2*cm,
            topMargin=2*cm, bottomMargin=2*cm
        )

        styles = getSampleStyleSheet()
        elements = []

        def style(name, **kwargs):
            return ParagraphStyle(name, parent=styles["Normal"], **kwargs)

        # Header
        header_style = style("Header", fontSize=20, textColor=white, alignment=TA_CENTER,
                              fontName="Helvetica-Bold", spaceAfter=4)
        sub_style = style("Sub", fontSize=10, textColor=HexColor("#94A3B8"), alignment=TA_CENTER)

        header_table = Table([[
            Paragraph("🏥 CIOS Clinical Intelligence Report", header_style),
        ]], colWidths=[17*cm])
        header_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), NAVY),
            ("TOPPADDING", (0, 0), (-1, -1), 16),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 16),
            ("ROUNDEDCORNERS", [6]),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 0.3*cm))
        elements.append(Paragraph(
            f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | By: {generated_by}",
            sub_style
        ))
        elements.append(Spacer(1, 0.5*cm))

        # Risk Level Banner
        risk = ai_assessment.get("risk_assessment", {})
        risk_level = risk.get("risk_level", "stable").upper()
        risk_score = risk.get("risk_score", 0)
        risk_color = {"CRITICAL": CRITICAL, "MODERATE": HIGH, "STABLE": LOW}.get(risk_level, TEAL)

        risk_banner = Table([[
            Paragraph(f"AI RISK LEVEL: {risk_level} — Score: {risk_score:.2%} | Confidence: {risk.get('confidence_score', 0):.0%}",
                      style("Risk", fontSize=13, textColor=white, fontName="Helvetica-Bold", alignment=TA_CENTER))
        ]], colWidths=[17*cm])
        risk_banner.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), risk_color),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ]))
        elements.append(risk_banner)
        elements.append(Spacer(1, 0.5*cm))

        # Patient Information
        section_style = style("Section", fontSize=12, fontName="Helvetica-Bold", textColor=NAVY, spaceAfter=8)
        label_style = style("Label", fontSize=9, textColor=HexColor("#64748B"), fontName="Helvetica-Bold")
        value_style = style("Value", fontSize=10, textColor=black)

        elements.append(Paragraph("PATIENT INFORMATION", section_style))
        elements.append(HRFlowable(width="100%", thickness=1, color=TEAL))
        elements.append(Spacer(1, 0.2*cm))

        dob = patient.get("date_of_birth", "")
        age = ""
        if dob:
            try:
                d = datetime.fromisoformat(str(dob).replace("Z", ""))
                age = f"{int((datetime.utcnow() - d).days / 365)} years"
            except:
                pass

        pt_data = [
            ["Full Name", patient.get("full_name", "—"), "Patient ID", patient.get("patient_id", "—")],
            ["Date of Birth", f"{dob} ({age})", "Gender", patient.get("gender", "—")],
            ["Blood Type", patient.get("blood_type", "—"), "Status", patient.get("status", "—").upper()],
            ["Ward", patient.get("ward", "—"), "Bed", patient.get("bed_number", "—")],
            ["Allergies", ", ".join(patient.get("allergies", [])) or "None", "Medications",
             str(len(patient.get("current_medications", []))) + " active"],
        ]

        pt_table = Table(pt_data, colWidths=[4*cm, 5.5*cm, 4*cm, 5.5*cm])
        pt_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), LIGHT_BG),
            ("BACKGROUND", (2, 0), (2, -1), LIGHT_BG),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#E2E8F0")),
            ("PADDING", (0, 0), (-1, -1), 6),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(pt_table)
        elements.append(Spacer(1, 0.5*cm))

        # AI Assessment
        elements.append(Paragraph("AI CLINICAL ASSESSMENT", section_style))
        elements.append(HRFlowable(width="100%", thickness=1, color=TEAL))
        elements.append(Spacer(1, 0.2*cm))

        summary = ai_assessment.get("clinical_summary", "No AI summary available.")
        elements.append(Paragraph(summary, style("Summary", fontSize=10, leading=16, spaceAfter=10,
                                                   backColor=HexColor("#EFF6FF"),
                                                   borderPadding=(8, 8, 8, 8))))
        elements.append(Spacer(1, 0.2*cm))

        # Contributing factors
        if risk.get("contributing_factors"):
            elements.append(Paragraph("Risk Contributing Factors:", label_style))
            for factor, weight in list(risk.get("contributing_factors", {}).items())[:6]:
                bar_width = int(weight * 30)
                bar = "█" * max(1, bar_width)
                elements.append(Paragraph(
                    f"• {factor.replace('_', ' ').title()}: {bar} {weight:.3f}",
                    style("Factor", fontSize=9, fontName="Courier", leftIndent=10)
                ))
            elements.append(Spacer(1, 0.3*cm))

        # Recommendations
        recs = risk.get("recommendations", [])
        if recs:
            elements.append(Paragraph("Recommendations:", label_style))
            for rec in recs:
                elements.append(Paragraph(f"→ {rec}", style("Rec", fontSize=9, leftIndent=10, spaceAfter=3)))
            elements.append(Spacer(1, 0.3*cm))

        # Vitals
        if vitals:
            elements.append(Paragraph("RECENT VITAL SIGNS", section_style))
            elements.append(HRFlowable(width="100%", thickness=1, color=TEAL))
            elements.append(Spacer(1, 0.2*cm))

            vital_headers = ["Time", "Temp(°C)", "HR(bpm)", "BP(mmHg)", "SpO2(%)", "RR", "GCS"]
            vital_rows = [vital_headers]
            for v in vitals[-5:]:
                ts = v.get("recorded_at", "")
                if isinstance(ts, str) and "T" in ts:
                    ts = ts.split("T")[1][:5]
                vital_rows.append([
                    str(ts),
                    str(v.get("temperature", "—")),
                    str(v.get("heart_rate", "—")),
                    f"{v.get('systolic_bp', '?')}/{v.get('diastolic_bp', '?')}",
                    str(v.get("oxygen_saturation", "—")),
                    str(v.get("respiratory_rate", "—")),
                    str(v.get("gcs_score", "—")),
                ])

            v_table = Table(vital_rows, colWidths=[2.8*cm, 2.4*cm, 2.4*cm, 2.8*cm, 2.4*cm, 1.8*cm, 1.8*cm])
            v_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), NAVY),
                ("TEXTCOLOR", (0, 0), (-1, 0), white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#E2E8F0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, LIGHT_BG]),
                ("PADDING", (0, 0), (-1, -1), 5),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ]))
            elements.append(v_table)
            elements.append(Spacer(1, 0.5*cm))

        # Diagnoses
        if diagnoses:
            elements.append(Paragraph("ACTIVE DIAGNOSES", section_style))
            elements.append(HRFlowable(width="100%", thickness=1, color=TEAL))
            elements.append(Spacer(1, 0.2*cm))

            for dx in diagnoses[:8]:
                sev = dx.get("severity", "unknown")
                sev_color = {"critical": CRITICAL, "severe": HIGH, "moderate": MEDIUM, "mild": LOW}.get(sev, TEAL)
                dx_line = Table([[
                    Paragraph(f"[{dx.get('icd_code', 'N/A')}] {dx.get('condition_name', '—')}",
                              style("DX", fontSize=9, fontName="Helvetica-Bold")),
                    Paragraph(sev.upper(), style("Sev", fontSize=8, textColor=white, alignment=TA_CENTER)),
                ]], colWidths=[14*cm, 3*cm])
                dx_line.setStyle(TableStyle([
                    ("BACKGROUND", (1, 0), (1, 0), sev_color),
                    ("PADDING", (0, 0), (-1, -1), 5),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]))
                elements.append(dx_line)
                elements.append(Spacer(1, 0.1*cm))

        # Footer
        elements.append(Spacer(1, 1*cm))
        footer = Table([[
            Paragraph(
                f"CIOS Clinical Intelligence OS | Confidential Medical Document | "
                f"Report ID: {filename} | This report contains AI-generated analysis. "
                f"Clinical decisions must be validated by qualified medical professionals.",
                style("Footer", fontSize=7, textColor=HexColor("#94A3B8"), alignment=TA_CENTER)
            )
        ]], colWidths=[17*cm])
        footer.setStyle(TableStyle([
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("LINEABOVE", (0, 0), (-1, 0), 0.5, HexColor("#CBD5E1")),
        ]))
        elements.append(footer)

        doc.build(elements)
        logger.info(f"[Reports] PDF generated: {filepath}")
        return filepath

    def _generate_csv(self, filename, patient, vitals, diagnoses, labs, ai_assessment) -> str:
        filepath = os.path.join(settings.REPORTS_DIR, f"{filename}.csv")
        risk = ai_assessment.get("risk_assessment", {})

        with open(filepath, "w", newline="") as f:
            writer = csv.writer(f)

            writer.writerow(["CIOS Clinical Report", datetime.utcnow().isoformat()])
            writer.writerow([])

            writer.writerow(["=== PATIENT INFO ==="])
            for k, v in patient.items():
                writer.writerow([k, str(v)])
            writer.writerow([])

            writer.writerow(["=== AI ASSESSMENT ==="])
            writer.writerow(["risk_score", risk.get("risk_score")])
            writer.writerow(["risk_level", risk.get("risk_level")])
            writer.writerow(["confidence_score", risk.get("confidence_score")])
            writer.writerow(["explanation", "; ".join(risk.get("explanation", []))])
            writer.writerow([])

            writer.writerow(["=== VITAL SIGNS ==="])
            if vitals:
                writer.writerow(list(vitals[0].keys()))
                for v in vitals:
                    writer.writerow(list(v.values()))
            writer.writerow([])

            writer.writerow(["=== DIAGNOSES ==="])
            if diagnoses:
                writer.writerow(["condition", "icd_code", "severity", "diagnosed_at"])
                for dx in diagnoses:
                    writer.writerow([dx.get("condition_name"), dx.get("icd_code"),
                                     dx.get("severity"), dx.get("diagnosed_at")])

        return filepath

    def _generate_excel(self, filename, patient, vitals, diagnoses, labs, ai_assessment) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return self._generate_csv(filename, patient, vitals, diagnoses, labs, ai_assessment)

        filepath = os.path.join(settings.REPORTS_DIR, f"{filename}.xlsx")
        wb = openpyxl.Workbook()
        risk = ai_assessment.get("risk_assessment", {})

        # Style helpers
        NAVY_FILL = PatternFill(start_color="0B1E3D", end_color="0B1E3D", fill_type="solid")
        TEAL_FILL = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
        HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
        BOLD = Font(bold=True)

        # Sheet 1: Summary
        ws = wb.active
        ws.title = "Summary"
        ws.append(["CIOS Clinical Intelligence Report", "", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")])
        ws["A1"].font = Font(bold=True, size=14, color="0B1E3D")
        ws.append([])
        ws.append(["AI Risk Score", risk.get("risk_score", 0),
                   "Risk Level", risk.get("risk_level", "").upper(),
                   "Confidence", risk.get("confidence_score", 0)])
        for k, v in patient.items():
            ws.append([str(k).replace("_", " ").title(), str(v)])

        # Sheet 2: Vitals
        if vitals:
            ws2 = wb.create_sheet("Vital Signs")
            headers = list(vitals[0].keys())
            ws2.append(headers)
            for cell in ws2[1]:
                cell.fill = NAVY_FILL
                cell.font = HEADER_FONT
            for v in vitals:
                ws2.append(list(v.values()))
            for i, col in enumerate(ws2.columns, 1):
                ws2.column_dimensions[get_column_letter(i)].width = 16

        # Sheet 3: Diagnoses
        if diagnoses:
            ws3 = wb.create_sheet("Diagnoses")
            ws3.append(["Condition", "ICD Code", "Severity", "Treatment Plan", "Diagnosed At"])
            for cell in ws3[1]:
                cell.fill = TEAL_FILL
                cell.font = HEADER_FONT
            for dx in diagnoses:
                ws3.append([dx.get("condition_name"), dx.get("icd_code"),
                             dx.get("severity"), dx.get("treatment_plan"), dx.get("diagnosed_at")])

        # Sheet 4: AI Analysis
        ws4 = wb.create_sheet("AI Analysis")
        ws4.append(["AI Clinical Intelligence Analysis"])
        ws4["A1"].font = Font(bold=True, size=13)
        ws4.append([])
        ws4.append(["Summary", ai_assessment.get("clinical_summary", "N/A")])
        ws4.append(["Risk Score", risk.get("risk_score")])
        ws4.append(["Risk Level", risk.get("risk_level")])
        ws4.append(["Confidence", risk.get("confidence_score")])
        ws4.append(["Requires Review", risk.get("requires_human_review")])
        ws4.append([])
        ws4.append(["Contributing Factors"])
        for k, v in (risk.get("contributing_factors") or {}).items():
            ws4.append([k.replace("_", " ").title(), v])
        ws4.append([])
        ws4.append(["Recommendations"])
        for rec in (risk.get("recommendations") or []):
            ws4.append([rec])

        wb.save(filepath)
        return filepath
